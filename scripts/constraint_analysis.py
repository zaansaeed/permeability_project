# inverse_design.py
import os
from pathlib import Path
import joblib
import numpy as np
import pandas as pd
import ast
from openpyxl import Workbook
from openpyxl.styles import PatternFill
 


def here() -> str:
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# === 0) Paths ===
os.chdir(os.path.dirname(os.path.abspath(__file__)))
folder = here()

MODEL_PATH = folder + "/saved_model/random_forest_model.joblib"
X_PATH     = folder + "/saved_model/X.csv"
Y_PATH     = folder + "/saved_model/y.csv"
MONOMER_LIST_CSV = folder + "/data/monomer_list_updated.csv"


PEPTIDE_CSV = folder + "/data/processed_peptides.csv"
 
# Build set of monomers actually used in training data
_pep_df = pd.read_csv(PEPTIDE_CSV)
TRAINING_MONOMERS = set()
for seq in _pep_df["Sequence"]:
    TRAINING_MONOMERS.update(ast.literal_eval(seq))
print(f"[Info] Training monomers: {len(TRAINING_MONOMERS)} unique symbols")
 
 


# === 1) Load model and data ===
pipe = joblib.load(MODEL_PATH)
model = getattr(pipe, "named_steps", {}).get("model", pipe)

X = pd.read_csv(X_PATH)
FEATURES = list(X.columns)  # 18 columns: 6 LogP + 6 is_D + 6 is_NSub

N_POS = 6
LOGP_FEATURES     = [f"Pos_{i}_logP"       for i in range(1, N_POS + 1)]
CHIRALITY_FEATURES = [f"Pos_{i}_is_D"       for i in range(1, N_POS + 1)]
NSUB_FEATURES      = [f"Pos_{i}_is_NSub"    for i in range(1, N_POS + 1)]
BINARY_FEATURES    = CHIRALITY_FEATURES + NSUB_FEATURES

print(f"[Info] LogP features:     {LOGP_FEATURES}")
print(f"[Info] Chirality features: {CHIRALITY_FEATURES}")
print(f"[Info] NSub features:      {NSUB_FEATURES}")


def target_region(
    target=None,
    constraints=None,
    eps=0.05,
    n_samples=20000,
    jitter_frac=0.5,
    include_shap=False,
    case_study_name="CaseStudy1"
):
    rng = np.random.default_rng(0)

    base = X.sample(n=max(len(X), n_samples), replace=len(X) < n_samples, random_state=0).copy()

    # 1) LogP: Gaussian jitter
    logp_values = base[LOGP_FEATURES].values
    logp_std = X[LOGP_FEATURES].std().replace(0, 1e-9).values
    logp_jitter = rng.normal(0, 1, size=(len(base), N_POS)) * (jitter_frac * logp_std)
    jittered_logp = logp_values + logp_jitter

    # 2) Binary features: random 0/1
    random_chirality = rng.integers(0, 2, size=(len(base), N_POS))
    random_nmethyl   = rng.integers(0, 2, size=(len(base), N_POS))

    # 3) Combine
    cand = pd.DataFrame(
        np.hstack([jittered_logp, random_chirality, random_nmethyl]),
        columns=FEATURES
    )

    # Enforce constraints
    if constraints:
        for col, v in constraints.items():
            if col not in cand.columns:
                print(f"[target_region] Skipping constraint '{col}' — not in FEATURES.")
                continue
            if isinstance(v, (tuple, list)):
                lo, hi = float(v[0]), float(v[1])
                cand[col] = rng.uniform(lo, hi, size=len(cand))
            else:
                cand[col] = float(v)

    # Clip LogP to 1st–99th pct (unless constrained)
    ql, qh = X[LOGP_FEATURES].quantile(0.01), X[LOGP_FEATURES].quantile(0.99)
    for c in LOGP_FEATURES:
        if not constraints or c not in constraints:
            cand[c] = cand[c].clip(float(ql[c]), float(qh[c]))

    # Ensure binary features are 0 or 1
    for c in BINARY_FEATURES:
        cand[c] = cand[c].round().clip(0, 1).astype(int)

    # Predict & filter
    yhat = np.asarray(pipe.predict(cand[FEATURES])).ravel()
    cand["y_pred"] = yhat

    if target is not None:
        hit = cand.loc[np.abs(yhat - target) <= eps].copy()
    else:
        hit = cand.copy()

    if hit.empty:
        return None, {"msg": f"No candidates within ±{eps} of target. Try increasing eps, n_samples, or jitter_frac."}

    monomers = pd.read_csv(MONOMER_LIST_CSV)

    # === Output folder ===
    out_dir = os.path.join(here(), case_study_name)
    os.makedirs(out_dir, exist_ok=True)

    # === Save 0: Constraints ===
    constraints_rows = [
        {"parameter": "target",      "value": target},
        {"parameter": "eps",         "value": eps},
        {"parameter": "n_samples",   "value": n_samples},
        {"parameter": "jitter_frac", "value": jitter_frac},
    ]
    for k, v in (constraints or {}).items():
        constraints_rows.append({"parameter": k, "value": v})
    constraints_df = pd.DataFrame(constraints_rows)
    constraints_path = os.path.join(out_dir, "constraints.csv")
    constraints_df.to_csv(constraints_path, index=False)
    print(f"[Saved] Constraints → {constraints_path}")

    # === Save 1: All hits ===
    hits_df = hit[FEATURES + ["y_pred"]].copy()
    hits_path = os.path.join(out_dir, "all_hits.csv")
    hits_df.to_csv(hits_path, index=False)
    print(f"[Saved] {len(hits_df)} hits → {hits_path}")

    # === Save 2: LogP position summaries ===
    summary = pd.DataFrame({
        "min":    hit[LOGP_FEATURES].min(),
        "q25":    hit[LOGP_FEATURES].quantile(0.25),
        "median": hit[LOGP_FEATURES].median(),
        "q75":    hit[LOGP_FEATURES].quantile(0.75),
        "max":    hit[LOGP_FEATURES].max(),
    })
    summary.index.name = "feature"
    summary_path = os.path.join(out_dir, "logP_position_summary.csv")
    summary.to_csv(summary_path)
    print(f"[Saved] LogP summary → {summary_path}")

    # === Save 3: Possible monomers per position ===
    # Compute proportion of 1 for each binary feature across hits
    binary_proportions = {}
    for col in BINARY_FEATURES:
        binary_proportions[col] = float(hit[col].mean())

    
    monomer_lists = {}
    for pos in range(1, N_POS + 1):
        logp_col = f"Pos_{pos}_logP"
        q25, q75 = hit[logp_col].quantile(0.25), hit[logp_col].quantile(0.75)
        mask = (monomers["logP"] >= q25) & (monomers["logP"] <= q75)
        mask &= (monomers["Monomer_Type"] != "Terminal")  # exclude terminals
 
        # is_D filter
        d_col = f"Pos_{pos}_is_D"
        if constraints and d_col in constraints:
            mask &= (monomers["is_D"] == int(constraints[d_col]))
        else:
            prop_d = binary_proportions[d_col]
            if prop_d < 0.3:
                mask &= (monomers["is_D"] == 0)
            elif prop_d > 0.7:
                mask &= (monomers["is_D"] == 1)
 
        # is_NSub filter
        nm_col = f"Pos_{pos}_is_NSub"
        if constraints and nm_col in constraints:
            mask &= (monomers["is_NSub"] == int(constraints[nm_col]))
        else:
            prop_nm = binary_proportions[nm_col]
            if prop_nm < 0.3:
                mask &= (monomers["is_NSub"] == 0)
            elif prop_nm > 0.7:
                mask &= (monomers["is_NSub"] == 1)
 
        filtered = monomers.loc[mask, ["Symbol", "IUPAC_Name"]].dropna(subset=["Symbol"]).drop_duplicates(subset=["Symbol"])
        monomer_lists[logp_col] = {
            "Symbol": filtered["Symbol"].tolist(),
            "IUPAC":  filtered["IUPAC_Name"].tolist(),
        }
 
    # --- Write XLSX with green highlighting for training-set monomers ---
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    wb = Workbook()
    ws = wb.active
    ws.title = "Possible Monomers"
 
    # Header row
    headers = []
    for pos in range(1, N_POS + 1):
        pos_label = f"Pos_{pos}"
        d_col  = f"Pos_{pos}_is_D"
        nm_col = f"Pos_{pos}_is_NSub"
        headers.extend([f"{pos_label}_Symbol", f"{pos_label}_IUPAC"])
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
 
    # Row 2-3: proportion info
    for pos in range(1, N_POS + 1):
        col_offset = (pos - 1) * 2
        d_col  = f"Pos_{pos}_is_D"
        nm_col = f"Pos_{pos}_is_NSub"
        ws.cell(row=2, column=col_offset + 1, value=f"is_D prop={binary_proportions[d_col]:.3f}")
        ws.cell(row=3, column=col_offset + 1, value=f"is_NSub prop={binary_proportions[nm_col]:.3f}")
 
    # Data rows (starting row 4)
    max_len = max(len(v["Symbol"]) for v in monomer_lists.values()) if monomer_lists else 0
    for row_idx in range(max_len):
        for pos in range(1, N_POS + 1):
            logp_col = f"Pos_{pos}_logP"
            syms  = monomer_lists[logp_col]["Symbol"]
            iupac = monomer_lists[logp_col]["IUPAC"]
            col_offset = (pos - 1) * 2
 
            if row_idx < len(syms):
                sym_cell = ws.cell(row=row_idx + 4, column=col_offset + 1, value=syms[row_idx])
                iupac_cell = ws.cell(row=row_idx + 4, column=col_offset + 2,
                                     value=iupac[row_idx] if row_idx < len(iupac) else None)
                # Green if monomer is in training data
                if syms[row_idx] in TRAINING_MONOMERS:
                    sym_cell.fill = green_fill
                    iupac_cell.fill = green_fill
 
    # Auto-width columns
    for col in ws.columns:
        max_width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_width + 2, 40)
 
    monomers_path = os.path.join(out_dir, "possible_monomers_per_position.xlsx")
    wb.save(monomers_path)
    print(f"[Saved] Possible monomers → {monomers_path}")

    # Console preview
    def preview_list(lst, k=8):
        lst = list(lst)
        return lst if len(lst) <= k else lst[:k] + [f"...(+{len(lst)-k} more)"]

    summary["possible_monomers"] = [
        preview_list(monomer_lists[f"Pos_{pos}_logP"]["Symbol"], k=3)
        for pos in range(1, N_POS + 1)
    ]

    out = {
        "n_candidates": int(len(cand)),
        "n_hits":       int(len(hit)),
        "hit_rate":     float(len(hit) / len(cand)),
        "target":       target,
        "eps":          float(eps),
        "output_dir":   out_dir,
        "binary_proportions": binary_proportions,
    }

    shap_summary = None
    if include_shap:
        try:
            import shap
            tree_model = getattr(pipe, "named_steps", {}).get("model", model)
            explainer = shap.TreeExplainer(tree_model)
            phi = explainer.shap_values(hit[FEATURES].values)
            shap_summary = {"mean_abs_shap": dict(zip(FEATURES, np.mean(np.abs(phi), axis=0).tolist()))}
        except Exception as e:
            shap_summary = {"error": f"SHAP failed: {e}"}

    examples = hit.sample(min(10, len(hit)), random_state=1)
    examples = examples[FEATURES + ["y_pred"]]

    return {"feasible_ranges": summary, "examples": examples}, {**out, "shap": shap_summary}


# === Boxplot Generator ===
def generate_boxplot(hits_csv, out_dir):
    import matplotlib.pyplot as plt
    rng = np.random.default_rng(42)

    df = pd.read_csv(hits_csv)
    data   = [df[c].dropna().values for c in LOGP_FEATURES]
    labels = [c.replace("_logP", "").replace("_", " ") for c in LOGP_FEATURES]

    fig, ax = plt.subplots(figsize=(10, 6))
    bp = ax.boxplot(
        data, labels=labels, patch_artist=True, showmeans=True, showfliers=False,
        meanprops=dict(marker="D", markerfacecolor="red", markersize=5),
        boxprops=dict(facecolor="#AED6F1", edgecolor="black"),
        medianprops=dict(color="black", linewidth=1.5),
        whiskerprops=dict(color="black"),
        capprops=dict(color="black"),
    )

    for i, col_data in enumerate(data):
        q25, q75 = np.percentile(col_data, 25), np.percentile(col_data, 75)
        iqr = q75 - q25
        lo, hi = q25 - 1.5 * iqr, q75 + 1.5 * iqr
        outliers = col_data[(col_data < lo) | (col_data > hi)]
        jitter = rng.normal(0, 0.04, size=len(outliers))
        ax.scatter(np.full_like(outliers, i + 1) + jitter, outliers,
                   color="grey", s=15, alpha=0.6, zorder=3)
        whisker_hi = col_data[col_data <= hi].max() if any(col_data <= hi) else q75
        ax.text(i + 1, whisker_hi, f"n={len(outliers)}",
                ha="center", va="bottom", fontsize=7, fontweight="bold")

    all_vals = np.concatenate(data)
    margin = (all_vals.max() - all_vals.min()) * 0.1
    ax.set_ylim(all_vals.min() - margin, all_vals.max() + margin)
    ax.set_ylabel("LogP")
    ax.set_title("LogP Distribution per Position (Hits)")
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()

    plot_path = os.path.join(out_dir, "logP_boxplot.png")
    fig.savefig(plot_path, dpi=300)
    plt.close(fig)
    print(f"[Saved] Boxplot → {plot_path}")


# === Example usage ===
if __name__ == "__main__":
    TARGET = -5
    CASE_STUDY = "CaseStudy5"

    CONSTRAINTS = {
        CHIRALITY_FEATURES[0]: 0,
        LOGP_FEATURES[5]: [-1.85, 0.1],
        CHIRALITY_FEATURES[1]: 1,
        CHIRALITY_FEATURES[2]: 0,
        CHIRALITY_FEATURES[3]: 1,
        CHIRALITY_FEATURES[4]: 0,
        CHIRALITY_FEATURES[5]: 1,

    }

    y = pd.read_csv(Y_PATH).squeeze()

    res, meta = target_region(TARGET, constraints=CONSTRAINTS, eps=0.2, n_samples=20000,
                              include_shap=False, case_study_name=CASE_STUDY)

    if res is None:
        print(meta["msg"])
    else:
        print("=== Feasible ranges (hits only) - LogP features ===")
        print(res["feasible_ranges"].to_string())
        print("\n=== Example designs (first 10 sampled hits) ===")
        print(res["examples"].head(10).to_string(index=False))
        print(f"\nMETA: {meta}")
        print(f"\n[Output saved to: {meta['output_dir']}]")

        hits_csv = os.path.join(meta["output_dir"], "all_hits.csv")
        generate_boxplot(hits_csv, meta["output_dir"])